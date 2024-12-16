# Msc Ortho CFD python tkinter script File
# Computational Fluid Dynamics Lab
# Author = Uday Tummala
# Dr. Carlos F. Lange
# Department of Mechanical Engineering
# University of Alberta
#
# Date 2024-01-07
# cmd = python3.11 GUI_ortho.py

from tkinter import *
import os
import openpyxl
import subprocess
import shutil
import pandas as pd
import csv
from PIL import ImageTk,Image
from tkinter import messagebox
from tkinter import filedialog
from tkinter.filedialog import askdirectory
from tkinter import ttk

# Window Title, size

root = Tk()
root.title("Ortho CFD v0.1")
photo = ImageTk.PhotoImage(Image.open('CFDLab-blogo2.png'))
root.wm_iconphoto(False, photo)
root.geometry("600x800")
root.minsize(height=600, width=800)
root.maxsize(height=700, width=1000)

# Tab 1, Welcome page

def tab1():
    label1=Label(root, text='\n Univeristy of Alberta\n Ortho CFD Applicaton version 0.1\n\nWelcome', font=('Times_New_Roman', 25), bg = "green", bd = 50, fg = "white")
    label1.pack()
    label2=Label(root, text='\nThis App helps determine the pressure difference in the \nNasal cavity from 3D X-RAY Scan. This is done by using\nCFD analysis, the input file format must be in stl or obj.\n This App uses blender, openfoam and Paraview softwares.\n\n\nThis App is developed by University of Alberta.\nFor more information please contact Dr Carlos Lange\n email = clange@ualberta.ca', font=('Times_New_Roman', 15))
    label2.pack()
    button1=Button(root, text='Next', font=('Times_New_Roman',16), command=next_win1, activebackground='blue')
    button1.pack(side=BOTTOM)

# Tab 2, asks for information on patient and file details

def tab2():
    global un
    global pn
    global pa
    global pd
    global fn
    un = StringVar()
    pn = StringVar()
    pa = StringVar()
    pd = StringVar()
    fn = StringVar()
    un.set("")
    pn.set("")
    pa.set("")
    pd.set("")
    fn.set("")
    user_name = Label(root, text = "Username",font=('Times_New_Roman', 16)).place(x = 40, y = 60)
    patient_name = Label(root, text = "Patient Name",font=('Times_New_Roman', 16)).place(x = 40, y = 120)
    patient_Age = Label(root, text = "Patient Age",font=('Times_New_Roman', 16)).place(x = 40, y = 180)
    patient_Doctor = Label(root, text = "Doctor Name",font=('Times_New_Roman', 16)).place(x = 40, y = 240)
    user_name_e = Entry(root, width = 30, textvariable = un, font=('Times_New_Roman', 16)).place(x = 300, y = 60)
    patient_name_e = Entry(root, width = 30, textvariable = pn, font=('Times_New_Roman', 16)).place(x = 300,y = 120)
    patient_age_e = Spinbox(root,from_=1, to=120, textvariable = pa, width = 8,font=('Times_New_Roman', 16)).place(x = 300, y = 180)
    patient_doctor_e = Entry(root, width = 30, textvariable = pd, font=('Times_New_Roman', 16)).place(x = 300, y = 240)
    Label(root, text = "Type new or previous folder\nname to save files",font=('Times_New_Roman', 16)).place(x = 40, y = 330)
    fn_entry = Entry(root, width = 20, textvariable = fn, font=('Times_New_Roman', 16)).place(x = 400, y = 330)

    def open1():
        filepath = "gui_data.xlsx"
        if os.path.exists(filepath):
            top = Toplevel()
            top.title('Excel viewer')
            top.geometry("1200x400")
            top.minsize(height=400, width=1200)
            photo = ImageTk.PhotoImage(Image.open('CFDLab-blogo2.png'))
            top.wm_iconphoto(False, photo)
            wb1 = openpyxl.load_workbook(filepath)
            ws = wb1['tab1']
            max_row = ws.max_row
            max_col = ws.max_column
            global se
            se = IntVar()
            s_e = Spinbox(top,from_=1, to=max_row-1, width=6, textvariable = se, font=('Times_New_Roman', 14))
            s_e.pack()
            lb = Label(top, text = ("Which data to display of total", max_row-1), font=('Times_New_Roman', 14))
            lb.pack()

            def sh_fun():
                dd.insert(INSERT, "\n")
                for i in range(1, max_col):
                    if ws.cell(row = se.get()+1, column = i):
                        c_ob = ws.cell(row = se.get()+1, column = i)
                        dd.insert(END, ("--",c_ob.value))

            sb1 = Button(top, text='Show', font=('Times_New_Roman',14), command = sh_fun, activebackground='blue')
            sb1.pack()
            dd = Text(top, height = 20, width = 130, font=('Times_New_Roman', 12), bg = "light yellow", wrap=WORD)
            for i in range(1, max_col):
                row_text = ws.cell(row = 1, column = i)
                dd.insert(INSERT, ("--",row_text.value))
                dd.tag_add("start", "1.00","1.9999")
                dd.tag_config("start", background= "black", foreground= "white")

            dd.pack()
        else:
            messagebox.showwarning("Error", "No Data file")

    Label(root, text = "Click here to see previous\nfolder data",font=('Times_New_Roman', 16)).place(x = 40, y = 420)
    open_button = Button(root, text = "Open", command=open1, font=('Times_New_Roman', 16), activebackground='green', relief=GROOVE).place(x = 400, y = 420)

    for i in range(3):
        root.columnconfigure(i, weight=1)

    root.rowconfigure(1, weight=1)
    next_button = Button(root, text = "Next", font=('Times_New_Roman', 16), command =next_win2, activebackground='blue').grid(row=2, column=2, sticky='e')
    back_button = Button(root, text = "Back", font=('Times_New_Roman', 16), command =back_win2, activebackground='blue').grid(row=2, column=0, sticky='w')


# Tab 3, asks for Geometry location, and runs blender script and shows image

def tab3():
    global spp
    spp = StringVar()
    spp.set("")
    global file_path_var
    file_path_var = StringVar()
    file_path_var.set("")
    loc_label = Label(root, text = "Please open 3D-model input file",font=('Times_New_Roman', 16)).place(x = 40, y = 60)

    def open1():
        root_filename = filedialog.askopenfilename(initialdir="", title="select a file", filetypes=(("stl files", "*.stl"),("obj files", "*.obj")))
        loc_Label = Label(root, text=root_filename, font=('Times_New_Roman', 16),fg="white", bg="blue").place(x = 40, y = 100)
        file_path_var.set(root_filename)
        file_path_var.get()
        path1 = os.path.join(dirs, "geo_in.txt")
        file1 = open(path1,"w")
        file1.write(root_filename)
        file1.close()

    def run1():
        status_e.delete(0, END)
        status_e.insert(0, "")
        source_dir = r"data/Master_cfd_file"
        shutil.copytree(source_dir, dirs, symlinks=False, ignore=None, ignore_dangling_symlinks=False, dirs_exist_ok=True)
        proc = subprocess.Popen(["blender --background --python blender_ortho.py"], stdout = subprocess.PIPE, shell=True)
        (out, err) = proc.communicate()
        if b"Finished" in out:
            status_e.insert("insert","completed")
        else:
            status_e.insert("insert","error")

    open_button = Button(root, text="Open", command=open1, activebackground='green',	font=('Times_New_Roman', 16)).place(x = 400, y = 60)
    Run_Pre = Label(root, text = "Run Pre-Processing of Geometry",font=('Times_New_Roman', 16)).place(x = 40, y = 160)
    Run_button = Button(root, text = "Run", command=run1,font=('Times_New_Roman', 16), activebackground='green', relief=GROOVE).place(x = 400, y = 160)
    status = Label(root, text = "Status of Geometry Pre-Processing", font=('Times_New_Roman', 16)).place(x = 40, y = 240)
    progressbar = ttk.Progressbar(mode="indeterminate")
    progressbar.start()
    progressbar.place(x = 40, y = 280, height=28, width=440)
    progressbar.step(25)
    status_e = Entry(root, width = 30, borderwidth=4, textvariable = spp, font=('Times_New_Roman', 14),bg="blue", fg="white")
    status_e.place(x = 40, y = 320)
    Click = Label(root, text = "Click here to see geometry", font=('Times_New_Roman', 16)).place(x = 40, y = 400)

    def open2():
        top = Toplevel()
        top.title('Geo IMG after Pre-proc')
        top.wm_iconphoto(False, photo)
        path = os.path.join(dirs, "blendout.jpg")
        my_img = ImageTk.PhotoImage(Image.open(path))
        panel = Label(top, image=my_img)
        panel.pack()
        panel.photo = my_img

    click_button = Button(root, text = "Open", font=('Times_New_Roman', 16), activebackground='green',command=open2,relief=GROOVE).place(x = 400, y = 400)
    global cvar
    cvar = IntVar()
    c = Checkbutton(root, text="Check this box if displayed image of geometry good", font=('Times_New_Roman', 16), variable=cvar, onvalue=1, offvalue=0)
    c.deselect()
    c.place(x = 40, y = 480)

    if os.path.exists(os.path.join(dirs, "geo_in.txt")) and os.path.exists(os.path.join(dirs, "blendout.jpg")):
        status_e.delete(0, END)
        status_e.insert(0, "completed")
        path101 = os.path.join(dirs, "geo_in.txt")
        file99 = open(path101, "r")
        vv1 = file99.readline()
        file99.close()
        loc_Label = Label(root, text=vv1, font=('Times_New_Roman', 16),fg="white", bg="blue").place(x = 40, y = 100)

    for i in range(3):
        root.columnconfigure(i, weight=1)

    root.rowconfigure(1, weight=1)
    next_button = Button(root, text = "Next", font=('Times_New_Roman', 16), command =next_win3, activebackground='blue').grid(row=2, column=2, sticky='e')
    back_button = Button(root, text = "Back", font=('Times_New_Roman', 16), command =back_win3, activebackground='blue').grid(row=2, column=0, sticky='w')

# Tab 4, runs openfoam and shows residuals progress

def tab4():
    global pvfr
    global scfd
    pvfr = StringVar()
    scfd = StringVar()
    pvfr.set("")
    scfd.set("")
    patient_VFR = Label(root, text = "Patient VFR\n (Volume Flow rate, LPM)",font=('Times_New_Roman', 16)).place(x = 40, y = 60)
    patient_VFR_e = Spinbox(root,from_=1, to=500, textvariable = pvfr, font=('Times_New_Roman', 16)).place(x = 300, y = 60)
    ##
    def run1():
        status_e.delete(0, END)
        status_e.insert(0, "")
        path0 = os.path.join(dirs, "0", "pvfr.txt")
        file1 = open(path0,"w")
        file1.write("vfr" + " " + pvfr.get() + ";" + "\n#inputMode merge")
        file1.close()
        path2 = os.path.join(dirs, "constant/triSurface")
        proc1 = subprocess.Popen(['/bin/bash', '-c',"bash ./Allclean"], cwd=dirs, stdout=subprocess.PIPE)
        proc1.wait()
        procc = subprocess.Popen(['/bin/bash', '-c',"rm combined.stl"], cwd=path2, stdout=subprocess.PIPE)
        procc.wait()
        proc2 = subprocess.Popen(["cat *.stl >> combined.stl"], cwd=path2, shell=True, stdout=subprocess.PIPE)
        proc2.wait()
        proc = subprocess.Popen(['/bin/bash', '-c',"bash ./Allrun"], cwd=dirs, stdout=subprocess.PIPE)
        (out, err) = proc.communicate()
        if b"Finished" in out:
            status_e.insert("insert","completed")
        else:
            status_e.insert("insert","error")

    Run_CFD = Label(root, text = "Run CFD of the Geometry",font=('Times_New_Roman', 16)).place(x = 40, y = 160)
    Run_button = Button(root, text = "Run", command=run1, font=('Times_New_Roman', 16), activebackground='green', relief=GROOVE).place(x = 400, y = 160)
    Click = Label(root, text = "Click here to see CFD progress \n graph",font=('Times_New_Roman', 16)).place(x = 40, y = 260)
    
    def open1():
        top = Toplevel()
        top.title('Image of CFD progress graph')
        def refresh1():
            path2 = os.path.join(dirs, "gnuplot")
            proc1 = subprocess.Popen(['/bin/bash', '-c',"bash ./gnuplot_residuals"], cwd=path2, stdout=subprocess.PIPE)
            proc1.wait()

        path2 = os.path.join(dirs, "gnuplot", "residual_plot.png")
        if os.path.exists(path2):
            top.wm_iconphoto(False, photo)
            my_img1 = ImageTk.PhotoImage(Image.open(path2))
            img_label = Label(top, image=my_img1)
            img_label.pack()
            img_label.photo = my_img1
        else:
            refresh1

        btn1 = Button(top, text="Refresh", command = refresh1, activebackground='green', font=('Times_New_Roman', 16)).pack()
    ##
    click_button = Button(root, text = "Open", font=('Times_New_Roman', 16), activebackground='green',command=open1,relief=GROOVE).place(x = 400, y = 260)
    status = Label(root, text = "Status of CFD Analysis",font=('Times_New_Roman', 16)).place(x = 40, y = 340)
    progressbar = ttk.Progressbar(mode="indeterminate")
    progressbar.start()
    progressbar.place(x = 40, y = 380, height=28, width=400)
    progressbar.step(25)
    status_e = Entry(root, width = 30, borderwidth=4, textvariable = scfd, font=('Times_New_Roman', 14),bg="blue", fg="white", )
    status_e.place(x = 40, y = 420)
    if os.path.exists(os.path.join(dirs, "260", "U")) and os.path.exists(os.path.join(dirs, "0", "pvfr.txt")):
            status_e.delete(0, END)
            status_e.insert(0, "completed")

    for i in range(3):
        root.columnconfigure(i, weight=1)

    root.rowconfigure(1, weight=1)
    next_button = Button(root, text = "Next", font=('Times_New_Roman', 16), command =next_win4, activebackground='blue').grid(row=2, column=2, sticky='e')
    back_button = Button(root, text = "Back", font=('Times_New_Roman', 16), command =back_win4, activebackground='blue').grid(row=2, column=0, sticky='w')

# Tab 5, runs paraview script and shows pressure, velocity plots and displays values of pressure drop and airway resistance

def tab5():
    def run1():
        status_e.delete(0, END)
        status_e.insert(0, "")
        path2 = os.path.join(dirs, "postProcessing/avgsurf_in/0/surfaceFieldValue.dat")
        path3 = os.path.join(dirs, "postProcessing/avgsurf_out/0/surfaceFieldValue.dat")
        with open(path2) as f:
            reader = csv.reader(f, delimiter="\t")
            for line in reader:
                print(line)

        p_in = float(line[1])

        with open(path3) as f:
            reader = csv.reader(f, delimiter="\t")
            for line in reader:
                print(line)

        p_out = float(line[1])
        p_drop_e.insert("insert",(p_in - p_out) / 1000)
        a_res_e.insert("insert",(p_in - p_out) / (1000 * float(pvfr.get())))
        proc = subprocess.Popen(["pvbatch paraview_ortho.py"], stdout=subprocess.PIPE, shell=True)
        (out, err) = proc.communicate()
        if b"Finished" in out:
            status_e.insert("insert","completed")
        else:
            status_e.insert("insert","error")

    Run_post = Label(root, text = "Run Post-Processing of the CFD",font=('Times_New_Roman', 16)).place(x = 40, y = 60)
    Run_button = Button(root, text = "Run", command = run1, font=('Times_New_Roman', 16), activebackground='green', relief=GROOVE).place(x = 400, y = 60)
    click_label = Label(root, text = "Click here to see pressure, velocity plots of the Geometry",font=('Times_New_Roman', 16)).place(x = 40, y = 160)
    ###
    def open1():
        top = Toplevel()
        top.title('pressure, velocity plots')
        top.wm_iconphoto(False, photo)
        path1 = os.path.join(dirs, "p_cut_1.png")
        path2 = os.path.join(dirs, "p_cut_2.png")
        path3 = os.path.join(dirs, "p_full_1.png")
        path4 = os.path.join(dirs, "p_full_2.png")
        path5 = os.path.join(dirs, "p_full_3.png")
        path6 = os.path.join(dirs, "p_full_4.png")
        path7 = os.path.join(dirs, "v_cut_1.png")
        path8 = os.path.join(dirs, "v_cut_2.png")
        path9 = os.path.join(dirs, "v_streamlines1.png")
        path10 = os.path.join(dirs, "v_streamlines2.png")
        my_img1 = ImageTk.PhotoImage(Image.open(path1))
        my_img2 = ImageTk.PhotoImage(Image.open(path2))
        my_img3 = ImageTk.PhotoImage(Image.open(path3))
        my_img4 = ImageTk.PhotoImage(Image.open(path4))
        my_img5 = ImageTk.PhotoImage(Image.open(path5))
        my_img6 = ImageTk.PhotoImage(Image.open(path6))
        my_img7 = ImageTk.PhotoImage(Image.open(path7))
        my_img8 = ImageTk.PhotoImage(Image.open(path8))
        my_img9 = ImageTk.PhotoImage(Image.open(path9))
        my_img10 = ImageTk.PhotoImage(Image.open(path10))
        image_list = [my_img1, my_img2, my_img3, my_img4, my_img5, my_img6, my_img7, my_img8, my_img9, my_img10]
        my_label = Label(top,image=my_img1)
        my_label.pack()
        ##
        def forward(image_number):
            global my_label
            global button_forward
            global button_back
            for widgets in top.winfo_children():
                widgets.destroy()

            my_label = Label(top,image=image_list[image_number-1])
            button_forward = Button(top, text=">>",font=('Times_New_Roman', 26), command=lambda: forward(image_number+1))
            button_back = Button(top, text="<<", font=('Times_New_Roman', 26),command=lambda: back(image_number-1))
            if image_number == 10:
                button_forward = Button(top, text=">>", state=DISABLED)

            my_label.pack()
            button_back.pack(side = LEFT)
            button_forward.pack(side = RIGHT)
        ##
        def back(image_number):
            global my_label
            global button_forward
            global button_back
            for widgets in top.winfo_children():
                widgets.destroy()

            my_label = Label(top,image=image_list[image_number-1])
            button_forward = Button(top, text=">>",font=('Times_New_Roman', 26), command=lambda: forward(image_number+1))
            button_back = Button(top, text="<<",font=('Times_New_Roman', 26),command=lambda: back(image_number-1))
            if image_number == 10:
                button_back = Button(top, text="<<",font=('Times_New_Roman', 26),state=DISABLED)

            my_label.pack()
            button_back.pack(side = LEFT)
            button_forward.pack(side = RIGHT)
        ##
        button_back = Button(top, text="<<",font=('Times_New_Roman', 26), command=back, state=DISABLED)
        button_forward = Button(top, text=">>",font=('Times_New_Roman', 26), command=lambda: forward(2))
        button_back.pack(side = LEFT)
        button_forward.pack(side = RIGHT)
    ##
    click_button = Button(root, text = "Open", font=('Times_New_Roman', 16), activebackground='green', command=open1, relief=GROOVE).place(x = 600, y = 160)
    status = Label(root, text = "Status of Post-processing",font=('Times_New_Roman', 16)).place(x = 40, y = 240)
    progressbar = ttk.Progressbar(mode="indeterminate")
    progressbar.start()
    progressbar.place(x = 40, y = 280, height=28, width=400)
    progressbar.step(25)
    global spop
    global pdr
    global arr
    spop = StringVar()
    pdr = StringVar()
    arr = StringVar()
    spop.set("")
    pdr.set("")
    arr.set("")
    status_e = Entry(root, width = 30, borderwidth=4, textvariable = spop, font=('Times_New_Roman', 16),bg="blue", fg="white")
    status_e.place(x = 40, y = 320)
    p_drop = Label(root, text = "Pressure drop (kPa)",font=('Times_New_Roman', 16)).place(x = 40, y = 400)
    p_drop_e = Entry(root, width = 15, textvariable = pdr, font=('Times_New_Roman', 16))
    p_drop_e.place(x = 300, y = 400)
    a_res = Label(root, text = "Air Resistance \n(cmH20/L/s)", font=('Times_New_Roman', 16)).place(x = 40, y = 440)
    a_res_e = Entry(root, width = 15, textvariable = arr, font=('Times_New_Roman', 16))
    a_res_e.place(x = 300, y = 440)
    Label(root, text = "Save above data", font=('Times_New_Roman', 16)).place(x = 40, y = 510)
    ##
    if os.path.exists(os.path.join(dirs, "postProcessing/avgsurf_in/0/surfaceFieldValue.dat")) and os.path.exists(os.path.join(dirs, "postProcessing/avgsurf_out/0/surfaceFieldValue.dat")):
        global p_in
        p_in = StringVar()
        p_in.set("")
        global p_out
        p_out = StringVar()
        p_out.set("")
        path201 = os.path.join(dirs, "postProcessing/avgsurf_in/0/surfaceFieldValue.dat")
        path301 = os.path.join(dirs, "postProcessing/avgsurf_out/0/surfaceFieldValue.dat")
        with open(path201) as f:
            reader = csv.reader(f, delimiter="\t")
            for line in reader:
                print(line)

        p_in = float(line[1])
        with open(path301) as f:
            reader = csv.reader(f, delimiter="\t")
            for line in reader:
                print(line)

        p_out = float(line[1])
        p_drop_e.delete(0, END)
        a_res_e.delete(0, END)
        p_drop_e.insert("insert",(p_in - p_out)/1000)
        a_res_e.insert("insert",(p_in - p_out) / (1000 * float(pvfr.get())))

    ##
    if os.path.exists(os.path.join(dirs, "p_cut_1.png")):
        if os.path.exists(os.path.join(dirs, "v_cut_2.png")):
            status_e.delete(0, END)
            status_e.insert(0, "completed")
    ##
    save_button = Button(root, text = "Save", font=('Times_New_Roman', 16), command=save_win5, activebackground='blue').place(x = 300, y = 510)
    back_button = Button(root, text = "Back", font=('Times_New_Roman', 16), command=back_win5, activebackground='blue')
    back_button.pack(side=BOTTOM)

# Next button on tab 1

def next_win1():
    for widgets in root.winfo_children():
        widgets.destroy()

    tab2()

# Back button on tab 2

def back_win2():
    for widgets in root.winfo_children():
        widgets.destroy()

    tab1()

# Next button on tab 2

def next_win2():
    if un.get() and pn.get() and pa.get() and pd.get() and fn.get():
        username = un.get()
        patientname = pn.get()
        patientage = pa.get()
        patientdoctor = pd.get()
        filename = fn.get()
        file1 = open("fn.txt","w")
        file1.write(filename)
        file1.close()
        if username and patientname and patientage and patientdoctor and filename:
            global path
            global dirs
            path = StringVar()
            dirs = StringVar()
            dirs.set("")
            path_ = askdirectory()
            path.set(path_)
            dirs = os.path.join(path.get(), fn.get())
            file1 = open("sdir.txt","w")
            file1.write(dirs)
            file1.close()
            if not os.path.exists(dirs):
                os.makedirs(dirs)
                messagebox.showinfo('Info','Folder name created successfully!')
            else:
                messagebox.showinfo("Info",'for you Info the folder name already exists')
            filepath = "gui_data.xlsx"
            if not os.path.exists(filepath):
                wb1 = openpyxl.Workbook()
                sh1 = wb1.create_sheet('tab1')
                heading = ["Date","User Name","Patient Name","Patient Age","Patient Doctor", "File Name", "File Location", "Status pre-proc","Geo chk","Patient VFR","Status CFD","Status Post-proc","Pressure drop (kPa)","Airway resistance (cmH20/L/S)"]
                sh1.append(heading)
                wb1.save(filepath)
                wb1 = openpyxl.load_workbook(filepath)
                ws = wb1['tab1']
                ws.append([os.popen('date').read(), username, patientname, patientage, patientdoctor, filename])
                wb1.save(filepath)
            else:
                wb1 = openpyxl.load_workbook(filepath)
                ws = wb1['tab1']
                ws.append([os.popen('date').read(), username, patientname, patientage, patientdoctor, filename])
                wb1.save(filepath)
        for widgets in root.winfo_children():
            widgets.destroy()

        tab3()

    else:
        messagebox.showerror("Error", "Missing information")

# Back button on tab 3

def back_win3():
    for widgets in root.winfo_children():
        widgets.destroy()

    tab2()

# Next button on tab 3

def next_win3():
    if cvar.get() == 1:
        if spp.get() == "completed":
            if os.path.exists(os.path.join(dirs, "geo_in.txt")) and os.path.exists(os.path.join(dirs, "blendout.jpg")):
                filepath = "gui_data.xlsx"
                wb1 = openpyxl.load_workbook(filepath)
                ws = wb1['tab1']
                ws.append([os.popen('date').read(), un.get(), pn.get(), pa.get(), pd.get(), fn.get(), file_path_var.get(), spp.get()])
                wb1.save(filepath)
                for widgets in root.winfo_children():
                    widgets.destroy()

                tab4()
            else:
                messagebox.showerror("Error", "Missing Geo Location")
        else:
            messagebox.showerror("Error", "Geo status should be completed")
    else:
        messagebox.showerror("Error", "Please check the image")

# Back button on tab 4

def back_win4():
    for widgets in root.winfo_children():
        widgets.destroy()

    tab3()

# Next button on tab 4

def next_win4():
    if scfd.get() == "completed":
        filepath = "gui_data.xlsx"
        wb1 = openpyxl.load_workbook(filepath)
        ws = wb1['tab1']
        ws.append([os.popen('date').read(), un.get(), pn.get(), pa.get(), pd.get(), fn.get(), file_path_var.get(), spp.get(), pvfr.get(), scfd.get()])
        wb1.save(filepath)
        for widgets in root.winfo_children():
            widgets.destroy()

        tab5()
    else:
        messagebox.showerror("Error", "CFD status should be completed")

# Back button on tab 5

def back_win5():
    for widgets in root.winfo_children():
        widgets.destroy()

    tab4()

# Next button on tab 5

def save_win5():
    if spop.get() == "completed":
        if pdr.get() and arr.get():
            filepath = "gui_data.xlsx"
            wb1 = openpyxl.load_workbook(filepath)
            ws = wb1['tab1']
            ws.append([os.popen('date').read(), un.get(), pn.get(), pa.get(), pd.get(), fn.get(), file_path_var.get(), spp.get(), pvfr.get(), scfd.get(), spop.get(), pdr.get(), arr.get()])
            wb1.save(filepath)
        else:
            messagebox.showerror("Error", "Values are Empty")
    else:
        messagebox.showerror("Error", "Please complete Post-processing to save")

tab1()

root.mainloop()
